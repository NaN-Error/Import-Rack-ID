import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
import os

def load_settings():
    if os.path.exists("source_file.txt"):
        with open("source_file.txt", "r") as file:
            source_file_path, source_sheet = file.read().split("\n")
            source_file_label.config(text=os.path.basename(source_file_path) + " - " + source_sheet)
            print("Loaded source file settings:", source_file_path, source_sheet)

    if os.path.exists("target_file.txt"):
        with open("target_file.txt", "r") as file:
            target_file_path, target_sheet = file.read().split("\n")
            target_file_label.config(text=os.path.basename(target_file_path) + " - " + target_sheet)
            print("Loaded target file settings:", target_file_path, target_sheet)

    if os.path.exists("source_file.txt") and os.path.exists("target_file.txt"):
        begin_button.config(state=tk.NORMAL)

def select_file(is_source):
    """
    Open a file dialog to select an Excel file and then choose a sheet from it.
    :param is_source: Boolean indicating whether the selected file is a source file.
    """
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        sheet_selection_window = tk.Toplevel(root)
        sheet_selection_window.title("Select Sheet")

        def onselect(evt):
            # Callback function for selecting a sheet
            w = evt.widget
            index = int(w.curselection()[0])
            value = w.get(index)
            save_file_path(file_path, value, is_source)
            sheet_selection_window.destroy()

        listbox = tk.Listbox(sheet_selection_window)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', onselect)
        for name in sheet_names:
            listbox.insert(tk.END, name)

def save_file_path(file_path, sheet_name, is_source):
    """
    Save the selected file path and sheet name to a text file.
    :param file_path: Path of the selected Excel file.
    :param sheet_name: Name of the selected sheet.
    :param is_source: Boolean indicating whether the file is a source file.
    """
    with open("source_file.txt" if is_source else "target_file.txt", "w") as file:
        file.write(file_path + "\n" + sheet_name)
    # Update UI labels and enable 'Begin' button if both files are selected
    if is_source:
        source_file_label.config(text=os.path.basename(file_path) + " - " + sheet_name)
    else:
        target_file_label.config(text=os.path.basename(file_path) + " - " + sheet_name)
    if os.path.exists("source_file.txt") and os.path.exists("target_file.txt"):
        begin_button.config(state=tk.NORMAL)
    load_settings()


def begin_process():
    """
    Begin processing the data by loading the Excel files and applying the specified logic.
    """
    try:
        # Load file paths from saved text files
        if os.path.exists("source_file.txt") and os.path.exists("target_file.txt"):
            with open("source_file.txt", "r") as file:
                source_file_path, source_sheet = file.read().split("\n")
            with open("target_file.txt", "r") as file:
                target_file_path, target_sheet = file.read().split("\n")
            
            # Load dataframes from Excel files
            source_df = pd.read_excel(source_file_path, sheet_name=source_sheet)
            target_df = pd.read_excel(target_file_path, sheet_name=target_sheet)
            print("Excel files loaded successfully.")

            # Check for required columns in source and target dataframes
            required_columns_source = ["Product ID", "Rack ID", "Duplicated", "Copied"]
            required_columns_target = ["Product ID", "Rack ID"]
            if not all(col in source_df.columns for col in required_columns_source):
                raise ValueError("Source file is missing required columns.")
            if not all(col in target_df.columns for col in required_columns_target):
                raise ValueError("Target file is missing required columns.")
            print("Required columns present in both source and target files.")

            # Process the source dataframe and get updates
            source_updates, target_updates = process_source_df(source_df, target_df)
            print(f"Processed {len(source_updates)} updates for source file and {len(target_updates)} for target file.")

            # Save updated Excel files
            save_updated_files(source_file_path, source_sheet, source_updates, target_file_path, target_sheet, target_updates)
            print("Updates saved to Excel files.")

            messagebox.showinfo("Success", "Data transfer completed successfully.")
        else:
            raise FileNotFoundError("Source or target file not found.")
    except Exception as e:
        messagebox.showerror("Error", str(e))
        print(f"Error: {e}")

def process_source_df(source_df, target_df):
    """
    Process the source dataframe by identifying duplicates, copying data to target,
    and updating the 'Copied' column.
    """
    # Convert columns to 'object' type to handle mixed data types
    source_df['Duplicated'] = source_df['Duplicated'].astype('object')
    source_df['Copied'] = source_df['Copied'].astype('object')
    target_df['Rack ID'] = target_df['Rack ID'].astype('object')

    source_updates = []
    target_updates = []

    # Identify and mark duplicates in the source dataframe
    duplicated_rows = source_df.duplicated(subset=['Product ID'], keep=False)
    source_df.loc[duplicated_rows, 'Duplicated'] = 'YES'

    # Prepare updates for duplicated rows
    for index, is_duplicated in enumerate(duplicated_rows):
        if is_duplicated:
            # Row index in Excel file is index + 2 (due to header and 1-based indexing)
            source_updates.append((index + 2, source_df.columns.get_loc('Duplicated') + 1, 'YES'))

    # Process each product in the source dataframe
    for index, row in source_df.iterrows():
        if row['Duplicated'] == 'YES' or row['Copied'] == 'YES':
            continue

        product_id = row['Product ID']
        rack_id = row['Rack ID']

        # Check for missing Product ID or Rack ID
        if pd.isna(product_id) or pd.isna(rack_id):
            continue

        # Find matching product in target dataframe and copy Rack ID if empty
        target_row_index = target_df[target_df['Product ID'] == product_id].index
        if not target_row_index.empty and pd.isna(target_df.at[target_row_index[0], 'Rack ID']):
            target_updates.append((target_row_index[0] + 2, target_df.columns.get_loc('Rack ID') + 1, str(rack_id)))
            source_updates.append((index + 2, source_df.columns.get_loc('Copied') + 1, 'YES'))

    return source_updates, target_updates


def save_updated_files(source_file_path, source_sheet, source_updates, target_file_path, target_sheet, target_updates):
    """
    Save the updated dataframes back to their respective Excel files.
    """
    update_excel_cells(source_file_path, source_sheet, source_updates)
    update_excel_cells(target_file_path, target_sheet, target_updates)

def update_excel_cells(file_path, sheet_name, updates):
    """
    Update specific cells in an Excel file without losing formatting.
    :param file_path: Path to the Excel file
    :param sheet_name: Name of the sheet to update
    :param updates: List of tuples (row, column, value) specifying the updates
    """
    print("Saving...")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    for row, col, value in updates:
        sheet.cell(row=row, column=col, value=value)

    wb.save(file_path)

# Get the directory of the script
script_dir = os.path.dirname(os.path.realpath(__file__))

# Set the current working directory to the script's directory
os.chdir(script_dir)

# Tkinter interface setup
root = tk.Tk()
root.title("Excel Data Transfer")

source_file_label = tk.Label(root, text="No Source File Selected")
source_file_label.pack()
select_source_button = tk.Button(root, text="Choose Source Excel File", command=lambda: select_file(True))
select_source_button.pack()

target_file_label = tk.Label(root, text="No Target File Selected")
target_file_label.pack()
select_target_button = tk.Button(root, text="Choose Target Excel File", command=lambda: select_file(False))
select_target_button.pack()

begin_button = tk.Button(root, text="Begin", state=tk.DISABLED, command=begin_process)
begin_button.pack()

print("Current Working Directory:", os.getcwd())
load_settings()

root.mainloop()
